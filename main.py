import sys
import os
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QLabel, QFileDialog, QComboBox, QTableWidget, QTableWidgetItem, QCheckBox
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sqlite3
import pandas as pd
from datetime import datetime

class ReviewApp(QWidget):
    def __init__(self):
        super().__init__()
        self.excel_file_path = None
        self.db_path = None
        self.setWindowTitle('Product Review Filter')
        self.setGeometry(100, 100, 600, 400)
        
        layout = QVBoxLayout()
        
        self.label = QLabel('Select a month to filter products needing review')
        layout.addWidget(self.label)
        
        self.import_button = QPushButton('Import Excel File')
        self.import_button.clicked.connect(self.import_file)
        layout.addWidget(self.import_button)
        
        self.month_combo = QComboBox()
        self.month_combo.addItems(['Select All', 'January', 'February', 'March', 'April', 'May', 'June', 
                                   'July', 'August', 'September', 'October', 'November', 'December'])
        self.month_combo.currentIndexChanged.connect(self.filter_data)
        layout.addWidget(self.month_combo)
        
        self.table = QTableWidget()
        layout.addWidget(self.table)
        self.setLayout(layout)
        
        self.set_current_month()
        self.clear_table()
        self.setup_database()

    def setup_database(self):
        app_data_path = os.path.join(os.getenv('LOCALAPPDATA'), 'ReviewApp')
        os.makedirs(app_data_path, exist_ok=True)

        self.db_path = os.path.join(app_data_path, 'product_reviews.db')

        if not os.path.exists(self.db_path):
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
            CREATE TABLE IF NOT EXISTS Products (
                product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_name TEXT NOT NULL,
                reference TEXT NOT NULL,
                review_date DATE NOT NULL,
                verified INTEGER NOT NULL
            )
            ''')

            conn.commit()
            conn.close()
            print("Database setup complete.")
        else:
            print("Database already exists. No setup needed.")

    def set_current_month(self):
        current_month_number = datetime.now().month
        self.month_combo.setCurrentIndex(current_month_number)

    def clear_table(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['Product Name', 'Reference', 'Review Date', 'Verified'])

    def detect_date_column(self, df):
        for col in df.columns:
            try:
                parsed_dates = pd.to_datetime(df[col], format='%d/%m/%Y', errors='coerce')
                if parsed_dates.notna().sum() / len(df) > 0.5:
                    return col
            except Exception as e:
                print(f"Error detecting date column: {e}")
        return None

    def import_excel_to_db(self, file_path):
        df = pd.read_excel(file_path)

        columns_map = self.detect_columns(df)

        if not columns_map:
            print("Required columns are missing in the Excel file.")
            return

        product_name_col = columns_map.get('product_name')
        reference_col = columns_map.get('reference')
        date_column = columns_map.get('date')
        verified_col = columns_map.get('verified')

        if not all([product_name_col, reference_col, date_column]):
            print("Error: Essential columns are missing in the Excel file.")
            return

        df[date_column] = pd.to_datetime(df[date_column], format='%d/%m/%Y', errors='coerce')

        if verified_col not in df.columns:
            df[verified_col] = False
        else:
            df[verified_col] = df[verified_col].fillna(0).astype(int)

        self.excel_file_path = file_path

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('DELETE FROM Products')

        for index, row in df.iterrows():
            review_date = row[date_column].strftime('%Y-%m-%d') if pd.notnull(row[date_column]) else None
            verified = row[verified_col] if verified_col in row else 0

            try:
                cursor.execute('''
                    INSERT INTO Products (product_name, reference, review_date, verified)
                    VALUES (?, ?, ?, ?)
                    ''', (row[product_name_col], row[reference_col], review_date, verified))
            except KeyError as e:
                print(f"Column not found: {e}")
            except Exception as e:
                print(f"An error occurred: {e}")

        conn.commit()
        conn.close()

        print(f"Data from {file_path} has been successfully imported.")

    def import_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)", options=options)
        if file_name:
            self.excel_file_path = file_name
            self.import_excel_to_db(file_name)
            self.filter_data()
        else:
            self.clear_table()

    def filter_data(self):
        if not self.excel_file_path:
            self.clear_table()
            return

        selected_month = self.month_combo.currentText()
        self.display_filtered_products(selected_month)

    def display_filtered_products(self, month):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        if month == 'Select All':
            cursor.execute('''
            SELECT product_name, reference, review_date, verified
            FROM Products
            ''')
        else:
            month_number = pd.to_datetime(month, format='%B').month
            cursor.execute('''
            SELECT product_name, reference, review_date, verified
            FROM Products
            WHERE strftime('%m', review_date) = ?
            ''', (f'{month_number:02d}',))

        records = cursor.fetchall()
        conn.close()

        self.table.setRowCount(len(records))
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['Product Name', 'Reference', 'Review Date', 'Verified'])

        for row_index, row_data in enumerate(records):
            for col_index, col_data in enumerate(row_data[:-1]):
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(col_data)))

            checkbox = QCheckBox()
            checkbox.setChecked(bool(row_data[3]))
            checkbox.stateChanged.connect(lambda state, ref=row_data[1]: self.update_verification(ref, state))
            self.table.setCellWidget(row_index, 3, checkbox)

    def update_verification(self, reference, state):
        verified = state == Qt.Checked

        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
            UPDATE Products
            SET verified = ?
            WHERE reference = ?
            ''', (verified, reference))

        self.update_excel_file(reference, verified)
        self.filter_data()

    def detect_columns(self, df):
        columns_map = {}
        
        df.columns = df.columns.str.strip()

        for col in df.columns:
            if df[col].dtype == object and df[col].str.len().max() > 5:
                columns_map['product_name'] = col
                break

        for col in df.columns:
            if df[col].dtype == object and df[col].str.match(r'^[A-Za-z0-9]+$').any():
                columns_map['reference'] = col
                break

        for col in df.columns:
            if df[col].dtype in [int, float] and df[col].dropna().isin([0, 1]).any():
                columns_map['verified'] = col
                break

        if 'date' not in columns_map:
            date_column = self.detect_date_column(df)
            if date_column:
                columns_map['date'] = date_column
            else:
                print("Error: No valid date column found.")
                return None

        return columns_map

    def update_excel_file(self, reference, verified):
        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            print("Excel file path not set or doesn't exist.")
            return

        try:
            # Read the Excel file into a DataFrame
            df = pd.read_excel(self.excel_file_path)

            # Clean column names to avoid issues like extra spaces
            df.columns = df.columns.str.strip()

            # Detect columns
            columns_map = self.detect_columns(df)
            
            if not columns_map:
                print("Error: Required columns are missing in the Excel file.")
                return

            # Extract column names using the map
            product_name_col = columns_map.get('product_name')
            reference_col = columns_map.get('reference')
            date_column = columns_map.get('date')
            verified_col = columns_map.get('verified')

            # Verify if essential columns are present
            if not all([product_name_col, reference_col, date_column, verified_col]):
                print("Error: Essential columns are missing in the Excel file.")
                return

            # Convert the detected date column to datetime format
            df[date_column] = pd.to_datetime(df[date_column], format='%d/%m/%Y', errors='coerce')

            # Update the 'Verified' status for the given reference
            df.loc[df[reference_col] == reference, verified_col] = int(verified)

            # Reformat the detected date column for saving to Excel
            df[date_column] = df[date_column].dt.strftime('%d/%m/%Y')

            # Save the updated DataFrame to the Excel file
            df.to_excel(self.excel_file_path, index=False)

            # Load the workbook to apply formatting
            wb = load_workbook(self.excel_file_path)
            ws = wb.active

            # Define colors for verified and not verified
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            # Apply colors to cells based on 'Verified' values
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(df.columns)):
                if row[len(df.columns) - 1].value == 1:  # Assuming 'Verified' is the last column
                    row[len(df.columns) - 1].fill = green_fill
                elif row[len(df.columns) - 1].value == 0:
                    row[len(df.columns) - 1].fill = red_fill

            # Save the workbook with formatting
            wb.save(self.excel_file_path)

            print(f"Excel file {self.excel_file_path} has been updated for reference {reference}.")

        except PermissionError as e:
            print(f"Permission error: {e}. Ensure the file is not open and has write permissions.")
        except Exception as e:
            print(f"An error occurred while updating the Excel file: {e}")



app = QApplication(sys.argv)
window = ReviewApp()
window.show()
sys.exit(app.exec_())
